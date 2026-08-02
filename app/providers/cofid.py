"""Normalizer for McCance and Widdowson's Composition of Foods Integrated Dataset.

CoFID is published by Public Health England / the Institute of Food Research
under the Open Government Licence v3.0, as a single spreadsheet whose nutrients
are split across sheets that all share one row order.

Layout of every nutrient sheet:

    row 1   long column heading with units
    row 2   short nutrient code (``PROT``, ``VITD``, ``NA``) — what we key on
    row 3   description
    row 4+  data, one row per food

Value conventions, quoted from the workbook's own notes sheet:

    ``N``       "present in significant quantities, but there is no reliable
                information on the amount" — unknown, so None
    ``Tr``      "a trace value" — measured and effectively zero, so 0
    blank       not measured — unknown, so None
    ``(1.2)``   parenthesised estimate — the value, 1.2

That distinction is why the nutrient columns are nullable: `N` and `Tr` are
different claims and must not both collapse to 0.

Nutrients are per 100 g of food except alcoholic beverages (group ``Q``), which
the publisher reports per 100 ml.
"""

from collections.abc import Iterator
from pathlib import Path

COFID_SOURCE = "cofid"

# Sheets carrying nutrients we map. The proximates sheet already includes the
# per-100g-food fatty acid totals and cholesterol, so the dedicated fatty acid
# sheets (1.7-1.12, mostly expressed per 100 g of fatty acids) are not needed.
NUTRIENT_SHEETS = ("1.3 Proximates", "1.4 Inorganics", "1.5 Vitamins")

CODE_ROW = 2
FIRST_DATA_ROW = 4
FOOD_CODE_COL = 0
NAME_COL = 1
GROUP_COL = 3

ALCOHOLIC_GROUP_PREFIX = "Q"

# Our field -> CoFID short codes, best first.
FIELD_CODES: dict[str, tuple[str, ...]] = {
    "calories_kcal": ("KCALS",),
    "protein_g": ("PROT",),
    "fat_g": ("FAT",),
    "carbs_g": ("CHO",),
    "sugar_g": ("TOTSUG",),
    # AOAC fibre is the modern method and comparable to USDA values; fall back
    # to non-starch polysaccharide for foods CoFID only measured the older way.
    "fiber_g": ("AOACFIB", "ENGFIB"),
    "saturated_fat_g": ("SATFOD",),
    "monounsaturated_fat_g": ("MONOFOD",),
    "polyunsaturated_fat_g": ("POLYFOD",),
    "trans_fat_g": ("FODTRANS",),
    "cholesterol_mg": ("CHOL",),
    "sodium_mg": ("NA",),
    "potassium_mg": ("K",),
    "calcium_mg": ("CA",),
    "magnesium_mg": ("MG",),
    "phosphorus_mg": ("P",),
    "iron_mg": ("FE",),
    "copper_mg": ("CU",),
    "zinc_mg": ("ZN",),
    "manganese_mg": ("MN",),
    "selenium_ug": ("SE",),
    "iodine_ug": ("I",),
    "vitamin_a_ug": ("RETEQU",),
    "vitamin_d_ug": ("VITD",),
    "vitamin_e_mg": ("VITE",),
    "vitamin_k_ug": ("VITK1",),
    "thiamin_mg": ("THIA",),
    "riboflavin_mg": ("RIBO",),
    "niacin_mg": ("NIAC",),
    "vitamin_b6_mg": ("VITB6",),
    "vitamin_b12_ug": ("VITB12",),
    "folate_ug": ("FOLT",),
    "pantothenic_acid_mg": ("PANTO",),
    "biotin_ug": ("BIOT",),
    "vitamin_c_mg": ("VITC",),
}

# CoFID does not report these at all; they stay unknown rather than zero.
UNREPORTED_FIELDS = (
    "added_sugar_g",
    "caffeine_mg",
    "chromium_ug",
    "choline_mg",
    "folic_acid_ug",
)

UNKNOWN_MARKER = "N"
TRACE_MARKER = "Tr"


def parse_cofid_value(raw) -> float | None:
    """Convert one CoFID cell into a nutrient value, or None when unknown."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if text == "" or text == UNKNOWN_MARKER:
        return None
    if text == TRACE_MARKER:
        return 0.0
    # Parenthesised values are the publisher's estimates, e.g. "(0.07)".
    if text.startswith("(") and text.endswith(")"):
        text = text[1:-1].strip()
    try:
        return float(text)
    except ValueError:
        return None


def normalize_cofid_food(
    values: dict[str, object],
    *,
    food_code: str,
    name: str,
    group: str | None,
) -> dict:
    """Build a food record from one food's CoFID short-code values."""
    nutrients: dict[str, float | None] = {}
    for field, codes in FIELD_CODES.items():
        value = None
        for code in codes:
            value = parse_cofid_value(values.get(code))
            if value is not None:
                break
        nutrients[field] = value
    for field in UNREPORTED_FIELDS:
        nutrients[field] = None

    is_alcoholic = bool(group) and group.upper().startswith(ALCOHOLIC_GROUP_PREFIX)

    return {
        "source": COFID_SOURCE,
        "source_code": food_code,
        "name": name,
        "brand": None,
        "barcode": None,
        "base_quantity": 100,
        "base_unit": "ml" if is_alcoholic else "g",
        "categories_tags": [f"cofid:{group}"] if group else [],
        **nutrients,
    }


def _sheet_codes(rows: list[tuple]) -> dict[int, str]:
    """Column index -> CoFID short nutrient code, from the header row."""
    header = rows[CODE_ROW - 1]
    codes = {}
    for index, cell in enumerate(header):
        code = str(cell).strip() if cell is not None else ""
        if code:
            codes[index] = code
    return codes


def read_cofid_workbook(path: str | Path) -> Iterator[dict]:
    """Yield normalized foods from a CoFID workbook.

    Sheets are joined by row position, which the publisher keeps aligned, and
    every row's food code is checked across sheets so a future release that
    reorders one sheet fails loudly instead of mixing nutrients between foods.
    """
    import openpyxl  # optional dependency, only needed for imports

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = {}
        for name in NUTRIENT_SHEETS:
            if name not in workbook.sheetnames:
                raise ValueError(
                    f"CoFID workbook is missing the {name!r} sheet; "
                    f"found {workbook.sheetnames}"
                )
            sheets[name] = list(workbook[name].iter_rows(values_only=True))

        primary = NUTRIENT_SHEETS[0]
        row_count = len(sheets[primary])
        for name, rows in sheets.items():
            if len(rows) != row_count:
                raise ValueError(
                    f"CoFID sheet {name!r} has {len(rows)} rows, "
                    f"expected {row_count} to match {primary!r}"
                )

        codes_by_sheet = {name: _sheet_codes(rows) for name, rows in sheets.items()}
        seen_codes: dict[str, int] = {}

        for index in range(FIRST_DATA_ROW - 1, row_count):
            base_row = sheets[primary][index]
            food_code = str(base_row[FOOD_CODE_COL] or "").strip()
            name_value = str(base_row[NAME_COL] or "").strip()
            if not food_code or not name_value:
                continue

            values: dict[str, object] = {}
            for sheet_name, rows in sheets.items():
                row = rows[index]
                row_code = str(row[FOOD_CODE_COL] or "").strip()
                if row_code != food_code:
                    raise ValueError(
                        f"CoFID sheet {sheet_name!r} row {index + 1} has food code "
                        f"{row_code!r}, expected {food_code!r} — sheets are not aligned"
                    )
                for column, code in codes_by_sheet[sheet_name].items():
                    if column < len(row):
                        values[code] = row[column]

            # CoFID 2021 reuses code 13-669 for two different foods. Keep both
            # by suffixing repeats, so neither silently upserts over the other.
            seen = seen_codes.get(food_code, 0) + 1
            seen_codes[food_code] = seen
            unique_code = food_code if seen == 1 else f"{food_code}#{seen}"

            group = base_row[GROUP_COL]
            yield normalize_cofid_food(
                values,
                food_code=unique_code,
                name=name_value,
                group=str(group).strip() if group else None,
            )
    finally:
        workbook.close()
