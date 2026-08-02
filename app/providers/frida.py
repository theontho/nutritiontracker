"""Normalizer for Frida, the Danish Food Composition Database.

Frida is published by the National Food Institute at DTU under CC BY 4.0. The
workbook carries the same data three ways; this reads `Data_Normalised`, which
is long-format with one row per measurement:

    FoodID | FoodName | ParameterID | ParameterName | ResVal | ...

As with CNF, a nutrient a food was never measured for simply has no row, so it
imports as None with nothing inferred from an absent value.

Nutrients are keyed by Frida's numeric ParameterID. The `Parameter` sheet gives
each one's published unit, which imports verify against the unit our field
expects, so a release that switches a nutrient between mg and ug fails rather
than silently scaling values by 1000.
"""

from collections.abc import Iterator
from pathlib import Path

FRIDA_SOURCE = "frida"

DATA_SHEET = "Data_Normalised"
PARAMETER_SHEET = "Parameter"
FOOD_SHEET = "Food"

# Our field -> Frida parameter IDs, best first.
FIELD_PARAMETERS: dict[str, tuple[str, ...]] = {
    "calories_kcal": ("356",),
    "protein_g": ("218",),
    "fat_g": ("141",),
    # Available carbohydrate is Frida's headline figure; fall back to
    # carbohydrate by difference where it is the only one given.
    "carbs_g": ("172", "170"),
    "sugar_g": ("245",),
    "added_sugar_g": ("417",),
    "fiber_g": ("168",),
    "saturated_fat_g": ("248",),
    "monounsaturated_fat_g": ("247",),
    "polyunsaturated_fat_g": ("251",),
    "trans_fat_g": ("261",),
    "cholesterol_mg": ("115",),
    "caffeine_mg": ("121",),
    "choline_mg": ("116",),
    "sodium_mg": ("201",),
    "potassium_mg": ("165",),
    "calcium_mg": ("108",),
    "magnesium_mg": ("184",),
    "phosphorus_mg": ("214",),
    "iron_mg": ("162",),
    "copper_mg": ("166",),
    "zinc_mg": ("274",),
    "manganese_mg": ("187",),
    "selenium_ug": ("230",),
    "iodine_ug": ("163",),
    "chromium_ug": ("117",),
    "vitamin_a_ug": ("12",),
    "vitamin_d_ug": ("126",),
    "vitamin_e_mg": ("135",),
    # Total vitamin K where available, otherwise K1 alone.
    "vitamin_k_ug": ("442", "164"),
    "thiamin_mg": ("37", "36"),
    "riboflavin_mg": ("39",),
    # Preformed niacin, matching how the other sources are mapped.
    "niacin_mg": ("294",),
    "vitamin_b6_mg": ("40",),
    "vitamin_b12_ug": ("38",),
    "folate_ug": ("143",),
    "pantothenic_acid_mg": ("210",),
    "biotin_ug": ("42",),
    "vitamin_c_mg": ("47",),
}

# Frida does not report these; they stay unknown rather than zero.
UNREPORTED_FIELDS = ("folic_acid_ug",)

# Frida writes units per 100 g and sometimes as an equivalence code, e.g.
# vitamin A as "RE (µg/100g)" and vitamin E as "alfa-TE" (mg of alpha
# tocopherol equivalents).
EQUIVALENCE_UNITS = {
    "alfa-te": "mg",
    "ne": "mg",
    "re": "ug",
}

FIELD_UNITS = {"g": "g", "mg": "mg", "ug": "ug", "kcal": "kcal"}


def parse_frida_value(raw) -> float | None:
    """Convert one Frida result value into a number, or None when absent."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if not text or text.upper() == "NULL":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_unit(unit: object) -> str:
    """Reduce a published unit string to the bare unit it measures in."""
    text = str(unit or "").strip().lower().replace("µ", "u")
    if not text:
        return ""
    equivalence = EQUIVALENCE_UNITS.get(text)
    if equivalence:
        return equivalence
    # "mg/100g", "kcal/100 g", "RE (ug/100g)" all reduce to the leading unit.
    for token in ("kcal", "kj", "mg", "ug", "g"):
        if f"{token}/100" in text:
            return token
    return text


def _field_unit(field: str) -> str:
    return field.rsplit("_", 1)[-1]


def require_published_unit(parameter_id: str, units_by_id: dict[str, str]) -> None:
    """Fail if a nutrient we import has no unit published in the metadata.

    `check_parameter_units` can only compare units it was handed, so a
    parameter missing from the Parameter sheet would import completely
    unverified — the exact silent rescale that checking units prevents.
    """
    if parameter_id not in units_by_id:
        raise ValueError(
            f"Frida parameter {parameter_id} has no unit in the "
            f"{PARAMETER_SHEET!r} sheet; refusing to import a value whose "
            "unit cannot be verified"
        )


def check_parameter_units(units_by_id: dict[str, str]) -> None:
    """Fail if Frida reports a mapped nutrient in a unit we do not expect."""
    for field, parameter_ids in FIELD_PARAMETERS.items():
        expected = FIELD_UNITS[_field_unit(field)]
        for parameter_id in parameter_ids:
            published = units_by_id.get(parameter_id)
            if published is None:
                continue
            actual = normalize_unit(published)
            if actual != expected:
                raise ValueError(
                    f"Frida parameter {parameter_id} for {field} is reported in "
                    f"{published!r} ({actual or 'unknown'}), expected {expected!r}"
                )


def normalize_frida_food(
    measurements: dict[str, float | None],
    *,
    food_id: str,
    name: str,
    group: str | None = None,
) -> dict:
    """Build a food record from one food's Frida parameter measurements."""
    nutrients: dict[str, float | None] = {}
    for field, parameter_ids in FIELD_PARAMETERS.items():
        value = None
        for parameter_id in parameter_ids:
            value = measurements.get(parameter_id)
            if value is not None:
                break
        nutrients[field] = value
    for field in UNREPORTED_FIELDS:
        nutrients[field] = None

    return {
        "source": FRIDA_SOURCE,
        "source_code": food_id,
        "name": name,
        "brand": None,
        "barcode": None,
        "base_quantity": 100,
        "base_unit": "g",
        "categories_tags": [f"frida:{group}"] if group else [],
        **nutrients,
    }


def _sheet_dicts(workbook, sheet_name: str) -> Iterator[dict]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Frida workbook is missing the {sheet_name!r} sheet; "
            f"found {workbook.sheetnames}"
        )
    rows = workbook[sheet_name].iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    for row in rows:
        yield dict(zip(header, row, strict=False))


def _wanted_parameters() -> set[str]:
    return {pid for ids in FIELD_PARAMETERS.values() for pid in ids}


def read_frida_workbook(path: str | Path) -> Iterator[dict]:
    """Yield normalized foods from a Frida dataset workbook."""
    import openpyxl  # optional dependency, only needed for imports

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        units_by_id = {
            str(row["ParameterID"]).strip(): row["Unit"]
            for row in _sheet_dicts(workbook, PARAMETER_SHEET)
            if row.get("ParameterID") is not None
        }
        check_parameter_units(units_by_id)

        groups_by_food = {
            str(row["FoodID"]).strip(): str(row.get("FoodGroupID") or "").strip()
            for row in _sheet_dicts(workbook, FOOD_SHEET)
            if row.get("FoodID") is not None
        }

        wanted = _wanted_parameters()
        names: dict[str, str] = {}
        order: list[str] = []
        measurements: dict[str, dict[str, float | None]] = {}

        for row in _sheet_dicts(workbook, DATA_SHEET):
            food_id = str(row.get("FoodID") or "").strip()
            if not food_id:
                continue
            if food_id not in names:
                name = str(row.get("FoodName") or "").strip()
                if not name:
                    continue
                names[food_id] = name
                order.append(food_id)
            parameter_id = str(row.get("ParameterID") or "").strip()
            if parameter_id not in wanted:
                continue
            require_published_unit(parameter_id, units_by_id)
            value = parse_frida_value(row.get("ResVal"))
            if value is None:
                continue
            measurements.setdefault(food_id, {})[parameter_id] = value

        for food_id in order:
            yield normalize_frida_food(
                measurements.get(food_id, {}),
                food_id=food_id,
                name=names[food_id],
                group=groups_by_food.get(food_id) or None,
            )
    finally:
        workbook.close()
