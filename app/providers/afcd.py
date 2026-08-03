"""Normalizer for the Australian Food Composition Database (AFCD).

AFCD is published by Food Standards Australia New Zealand under CC BY 4.0 as a
workbook with one wide row per food:

    row 1   title
    row 2   blank
    row 3   column headings, including the unit — what we key on
    row 4+  data, one row per food

Columns are matched by heading text rather than position, and every heading we
map must be present, so a re-ordered or renamed release fails loudly instead of
reading nutrients out of the wrong column.

Two traps in the published layout are handled here:

  - Fatty acid totals appear twice, once as a percentage of total fatty acids
    ("... (%T)") and once as grams per 100 g. Only the gram columns are used.
  - Total trans fatty acids are published in *milligrams* while the other fat
    columns are grams, so that column alone is scaled.

Energy is published in kilojoules only and is converted to kilocalories. A
blank cell means the food was never assayed for that nutrient, so it becomes
None rather than zero.
"""

import re
from collections.abc import Iterator
from pathlib import Path

AFCD_SOURCE = "afcd"

# The workbook also has a "Liquids only per 100 mL" sheet, but every food on it
# is already on the per-100 g sheet, so importing both would just duplicate 213
# foods under the same food keys.
NUTRIENT_SHEET = "All solids & liquids per 100 g"

HEADER_ROW = 3
FIRST_DATA_ROW = 4

FOOD_KEY_HEADER = "Public Food Key"
NAME_HEADER = "Food Name"
CLASSIFICATION_HEADER = "Classification"

KJ_PER_KCAL = 4.184

# Our field -> AFCD column heading, best first. Headings are matched after
# whitespace normalization.
FIELD_HEADERS: dict[str, tuple[str, ...]] = {
    "protein_g": ("Protein (g)",),
    "fat_g": ("Fat, total (g)",),
    # AFCD publishes available carbohydrate rather than carbohydrate by
    # difference; the "without sugar alcohols" figure is the one it uses in its
    # own energy calculation, so it stays consistent with the energy we import.
    "carbs_g": (
        "Available carbohydrate, without sugar alcohols (g)",
        "Available carbohydrate, with sugar alcohols (g)",
    ),
    "sugar_g": ("Total sugars (g)",),
    "added_sugar_g": ("Added sugars (g)",),
    "fiber_g": ("Total dietary fibre (g)",),
    "saturated_fat_g": ("Total saturated fatty acids, equated (g)",),
    "monounsaturated_fat_g": ("Total monounsaturated fatty acids, equated (g)",),
    "polyunsaturated_fat_g": ("Total polyunsaturated fatty acids, equated (g)",),
    "cholesterol_mg": ("Cholesterol (mg)",),
    "caffeine_mg": ("Caffeine (mg)",),
    "sodium_mg": ("Sodium (Na) (mg)",),
    "potassium_mg": ("Potassium (K) (mg)",),
    "calcium_mg": ("Calcium (Ca) (mg)",),
    "magnesium_mg": ("Magnesium (Mg) (mg)",),
    "phosphorus_mg": ("Phosphorus (P) (mg)",),
    "iron_mg": ("Iron (Fe) (mg)",),
    "copper_mg": ("Copper (Cu) (mg)",),
    "zinc_mg": ("Zinc (Zn) (mg)",),
    "manganese_mg": ("Manganese (Mn) (mg)",),
    "selenium_ug": ("Selenium (Se) (ug)",),
    "iodine_ug": ("Iodine (I) (ug)",),
    "chromium_ug": ("Chromium (Cr) (ug)",),
    "vitamin_a_ug": ("Vitamin A retinol equivalents (ug)",),
    "retinol_ug": ("Retinol (ug)",),
    "beta_carotene_ug": ("Beta-carotene (ug)",),
    "alpha_carotene_ug": ("Alpha-carotene (ug)",),
    "beta_cryptoxanthin_ug": ("Beta-cryptoxanthin (ug)",),
    "lycopene_ug": ("Lycopene (ug)",),
    "lutein_zeaxanthin_ug": ("Lutein and zeaxanthin (ug)",),
    "vitamin_d_ug": ("Vitamin D3 equivalents (ug)",),
    "vitamin_d2_ug": ("Vitamin D2 (ug)",),
    "vitamin_d3_ug": ("Vitamin D3 (ug)",),
    "vitamin_e_mg": ("Vitamin E (mg)",),
    "thiamin_mg": ("Thiamin (B1) (mg)",),
    "riboflavin_mg": ("Riboflavin (B2) (mg)",),
    "niacin_mg": ("Niacin (B3) (mg)",),
    "vitamin_b6_mg": ("Pyridoxine (B6) (mg)",),
    "vitamin_b12_ug": ("Cobalamin (B12) (ug)",),
    "folate_ug": ("Dietary folate equivalents (ug)", "Total folates (ug)"),
    "folic_acid_ug": ("Folic acid (ug)",),
    "pantothenic_acid_mg": ("Pantothenic acid (B5) (mg)",),
    "biotin_ug": ("Biotin (B7) (ug)",),
    "vitamin_c_mg": ("Vitamin C (mg)",),
}

ENERGY_HEADER = "Energy with dietary fibre, equated (kJ)"

# Published in milligrams while every other fat column is in grams.
TRANS_FAT_HEADER = "Total trans fatty acids, imputed (mg)"

# AFCD measures neither of these; they stay unknown rather than zero.
UNREPORTED_FIELDS = (
    "choline_mg",
    "vitamin_k_ug",
)


def normalize_header(text: object) -> str:
    """Collapse the workbook's irregular spacing so headings compare equal."""
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def parse_afcd_value(raw) -> float | None:
    """Convert one AFCD cell into a nutrient value, or None when unmeasured."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def kj_to_kcal(kilojoules: float | None) -> float | None:
    if kilojoules is None:
        return None
    return round(kilojoules / KJ_PER_KCAL, 1)


def normalize_afcd_food(
    values: dict[str, object],
    *,
    food_key: str,
    name: str,
    classification: str | None = None,
) -> dict:
    """Build a food record from one food's AFCD heading -> value mapping."""
    nutrients: dict[str, float | None] = {}
    for field, headers in FIELD_HEADERS.items():
        value = None
        for header in headers:
            value = parse_afcd_value(values.get(header))
            if value is not None:
                break
        nutrients[field] = value

    nutrients["calories_kcal"] = kj_to_kcal(parse_afcd_value(values.get(ENERGY_HEADER)))

    trans_mg = parse_afcd_value(values.get(TRANS_FAT_HEADER))
    nutrients["trans_fat_g"] = None if trans_mg is None else trans_mg / 1000

    for field in UNREPORTED_FIELDS:
        nutrients[field] = None

    return {
        "source": AFCD_SOURCE,
        "source_code": food_key,
        "name": name,
        "brand": None,
        "barcode": None,
        "base_quantity": 100,
        "base_unit": "g",
        "categories_tags": [f"afcd:{classification}"] if classification else [],
        **nutrients,
    }


def _expected_headers() -> set[str]:
    """Headings an import cannot proceed without.

    Only the first choice for each field is required: the later entries are
    fallbacks for foods the primary column does not cover, so a release that
    drops one should not block the import.
    """
    headers = {ENERGY_HEADER, TRANS_FAT_HEADER, FOOD_KEY_HEADER, NAME_HEADER}
    for options in FIELD_HEADERS.values():
        headers.add(options[0])
    return headers


def _header_columns(header_row: tuple) -> dict[str, int]:
    """Heading -> column index, checking every heading we map is present."""
    columns: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        heading = normalize_header(cell)
        if heading and heading not in columns:
            columns[heading] = index

    missing = sorted(h for h in _expected_headers() if h not in columns)
    if missing:
        raise ValueError(
            f"AFCD sheet is missing expected columns: {missing}. "
            "The release layout may have changed."
        )
    return columns


def read_afcd_workbook(path: str | Path) -> Iterator[dict]:
    """Yield normalized foods from an AFCD nutrient profiles workbook."""
    import openpyxl  # optional dependency, only needed for imports

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if NUTRIENT_SHEET not in workbook.sheetnames:
            raise ValueError(
                f"AFCD workbook is missing the {NUTRIENT_SHEET!r} sheet; "
                f"found {workbook.sheetnames}"
            )
        sheet = workbook[NUTRIENT_SHEET]

        columns: dict[str, int] | None = None
        key_column = name_column = classification_column = None

        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if index < HEADER_ROW:
                continue
            if index == HEADER_ROW:
                columns = _header_columns(row)
                key_column = columns[FOOD_KEY_HEADER]
                name_column = columns[NAME_HEADER]
                classification_column = columns.get(CLASSIFICATION_HEADER)
                continue
            if index < FIRST_DATA_ROW or columns is None:
                continue

            food_key = str(row[key_column] or "").strip()
            name = str(row[name_column] or "").strip()
            if not food_key or not name:
                continue

            values = {
                heading: row[column]
                for heading, column in columns.items()
                if column < len(row)
            }
            classification = None
            if classification_column is not None and classification_column < len(row):
                classification = str(row[classification_column] or "").strip() or None

            yield normalize_afcd_food(
                values,
                food_key=food_key,
                name=name,
                classification=classification,
            )
    finally:
        workbook.close()
